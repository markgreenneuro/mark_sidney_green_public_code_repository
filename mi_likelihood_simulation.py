#!/usr/bin/env python
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.backends.backend_pdf
from sklearn.feature_selection import SelectKBest, mutual_info_classif
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import pathlib
import shutil
import os
import sys
import time


def generate_data(n, feats_in_data):
    noise_matrix = pd.DataFrame(np.random.default_rng().uniform(0, 1, (n, feats_in_data)),
                                columns=['X' + str(x) for x in range(0, feats_in_data)])
    results_vector = pd.DataFrame(np.random.choice([1, 2, 3, 4, 5], noise_matrix.shape[0]), columns=['OUTCOME'])
    return [noise_matrix, results_vector]


def create_random_mapping():
    labels = np.array([1, 2, 3, 4, 5])
    np.random.shuffle(labels)
    labels_new = pd.DataFrame(labels, columns=['NEW_VALUE'])
    labels_old = pd.DataFrame([1, 2, 3, 4, 5], columns=['OLD_VALUE'])
    random_crossref_df = pd.concat([labels_old, labels_new], axis=1)
    return random_crossref_df


def create_additive_scaled_dataframe(results_vector, noise_matrix, scale, feats_in_data, random_crossref_df):
    additive_results_vector = np.array([])
    for j in range(results_vector.shape[0]):
        results_vector_instance = results_vector['OUTCOME'].iloc[j]
        for i in range(random_crossref_df.shape[0]):
            if results_vector_instance == random_crossref_df['OLD_VALUE'].iloc[i]:
                additive_results_vector = np.append(additive_results_vector, random_crossref_df['NEW_VALUE'].iloc[i])
    additive_results_vector = pd.DataFrame([additive_results_vector]).T
    additive_df = pd.concat([additive_results_vector] * feats_in_data, axis=1, ignore_index=True)
    additive_scaled_df = (additive_df / 5) * scale
    additive_scaled_df.columns = noise_matrix.columns
    return additive_scaled_df


def convert(input_set):
    output_list = input_set
    return [*output_list]


def create_substituted(feats_in_data, feats_to_alter, noise_matrix, results_vector, scale):
    random_crossref_df = create_random_mapping()
    additive_scaled_df = create_additive_scaled_dataframe(results_vector, noise_matrix, scale, feats_in_data,
                                                          random_crossref_df)
    selected_additive_scaled_df = additive_scaled_df.sample(feats_to_alter, axis='columns')
    set_columns_selected = set(selected_additive_scaled_df.columns)
    selected_noise = noise_matrix[convert(set_columns_selected)]
    set_columns_not_selected = set(noise_matrix.columns) - set(selected_additive_scaled_df.columns)
    not_selected_noise = noise_matrix[convert(set_columns_not_selected)]
    selected_noise_plus_additive = selected_noise + selected_additive_scaled_df / (1 + scale)
    substituted = pd.concat([not_selected_noise, selected_noise_plus_additive], axis=1)[noise_matrix.columns].dropna()
    return substituted


def create_edited_data(n, feats_in_data, scale, feat_num):
    [noise_matrix, results_vector] = generate_data(n, feats_in_data)
    substituted = create_substituted(feats_in_data, feat_num, noise_matrix, results_vector, scale)
    # substituted = create_substituted(feat_num, feats_in_data, noise_matrix, results_vector, scale)

    return [results_vector, substituted]


def select_features(substituted, results_vector, feat_num):
    fs = SelectKBest(score_func=mutual_info_classif, k=feat_num)
    fs.fit(substituted, np.array(results_vector).ravel())
    return fs


def calculate_mutual_information(fs, feats_train) -> pd.DataFrame:
    new_feats = pd.DataFrame(feats_train)[fs.get_support()].reset_index(drop=True)
    new_feats.columns = ['FEATS']
    mi = pd.DataFrame(fs.scores_)[fs.get_support()].reset_index(drop=True)
    mi.columns = ['MUTUAL_INFO']
    new_feats_mi = pd.concat([new_feats, mi], axis=1)
    [r, _] = new_feats_mi.shape
    num_feats = pd.DataFrame([np.repeat(str(r), r)]).T
    num_feats.columns = ['NUM_FEAT']
    new_feats_mi = pd.concat([num_feats, new_feats_mi], axis=1)
    return new_feats_mi


def get_params(new_feats_mi):
    num_params = new_feats_mi.shape[0]
    n = new_feats_mi.shape[1]
    return [num_params, n]


def find_mi_sum(new_feats_mi):
    mi_sum = float(np.sum(new_feats_mi['MUTUAL_INFO']))
    return mi_sum


def calculate_aic(n: int, mi_sum: float, num_params):
    if mi_sum != 0:
        aic = n * np.log(mi_sum) + 2 * num_params
    else:
        aic = 2 * num_params
    return aic


def calculate_bic(n: int, mi_sum: float, num_params: int) -> float:
    if mi_sum == 0:
        bic = num_params * np.log(n)
    else:
        bic = n * np.log(mi_sum) + num_params * np.log(n)
    return bic


def run_mutual_info_simulation(feats_to_alter, n, feats_in_data, scale):
    [results_vector, substituted] = create_edited_data(n, feats_in_data, scale, feats_to_alter)
    mic_array = np.empty([substituted.shape[1]])
    aic_array = np.empty([substituted.shape[1]])
    bic_array = np.empty([substituted.shape[1]])
    feat_num_array = np.empty([substituted.shape[1]])
    feats_train = substituted.columns
    for feat_num in range(1, substituted.shape[1] + 1):
        fs = select_features(substituted, results_vector, feat_num)
        new_feats_mi = calculate_mutual_information(fs, feats_train)
        mi_sum = find_mi_sum(new_feats_mi)
        [num_params, n] = get_params(new_feats_mi)
        aic = calculate_aic(n, mi_sum, num_params)
        bic = calculate_bic(n, mi_sum, num_params)
        mic_array[feat_num - 1] = np.array(mi_sum).astype(float)
        aic_array[feat_num - 1] = np.abs(aic).astype(float)
        bic_array[feat_num - 1] = np.abs(bic).astype(float)
        feat_num_array[feat_num - 1] = int(feat_num)
    return [mic_array, aic_array, bic_array, feat_num_array]


def loop_mutual_info_simulation(feats_in_data, n, scale):
    feat_mic_aic_bic = pd.DataFrame([])
    feat_num_array = np.array([])
    for i in range(feats_in_data):
        [mic_array, aic_array, bic_array, feat_num_array] = run_mutual_info_simulation(i, n, feats_in_data, scale)
        mic_array = pd.DataFrame(mic_array, columns=['MIC_' + str(i + 1)])
        aic_array = pd.DataFrame(aic_array, columns=['AIC_' + str(i + 1)])
        bic_array = pd.DataFrame(bic_array, columns=['BIC_' + str(i + 1)])
        feat_mic_aic_bic = pd.concat([feat_mic_aic_bic, mic_array, aic_array, bic_array], axis=1)

    feat_num_df = pd.DataFrame(feat_num_array, columns=['NUM_FEAT_ALTERED'])
    total_feat_df = pd.DataFrame(np.repeat(feat_mic_aic_bic.shape[0], [feats_in_data]), columns=['NUM_FEAT'])

    feat_mic_aic_bic = pd.concat([total_feat_df, feat_num_df, feat_mic_aic_bic], axis=1)
    feat_mic = pd.concat([total_feat_df, feat_num_df, feat_mic_aic_bic[feat_mic_aic_bic.filter(like='MIC').columns]],
                         axis=1)
    feat_aic = pd.concat([total_feat_df, feat_num_df, feat_mic_aic_bic[feat_mic_aic_bic.filter(like='AIC').columns]],
                         axis=1)
    feat_bic = pd.concat([total_feat_df, feat_num_df, feat_mic_aic_bic[feat_mic_aic_bic.filter(like='BIC').columns]],
                         axis=1)

    return [feat_mic_aic_bic, feat_mic, feat_aic, feat_bic]


def twist(feat_short):
    num_altered_feat_df = pd.DataFrame(feat_short['NUM_FEAT_ALTERED'])
    feat_alt_horiz_labels = pd.DataFrame(np.delete(np.array(feat_short.columns), [0, 1]),
                                         columns=['GOF_INDEX_NUM_FEAT_SELECT'])
    # A more recent python implementation would be:
    # feat_alt_horiz_labels = feat_alt_horiz_labels.merge(num_altered_feat_df, how="cross")
    key = np.array([int(1)])
    feat_alt_horiz_labels_key = pd.DataFrame(np.repeat(key, feat_alt_horiz_labels.shape[0]), columns=['KEY'])
    num_altered_feat_df_key = pd.DataFrame(np.repeat(key, num_altered_feat_df.shape[0]), columns=['KEY'])
    feat_alt_horiz_labels_key = pd.concat([feat_alt_horiz_labels_key, feat_alt_horiz_labels], axis=1)
    num_altered_feat_df_key = pd.concat([num_altered_feat_df_key, num_altered_feat_df], axis=1)
    feat_alt_horiz_labels = pd.merge(feat_alt_horiz_labels_key, num_altered_feat_df_key, on='KEY').drop(['KEY'], axis=1)

    feat_value_array = np.array([])
    for i in range(feat_alt_horiz_labels.shape[0]):
        gof_feat_label = feat_alt_horiz_labels.iloc[i][0]
        num_alt_feat = feat_alt_horiz_labels.iloc[i][1]
        feat_value = np.array(feat_short[feat_short['NUM_FEAT_ALTERED'] == num_alt_feat][gof_feat_label]).astype(float)
        feat_value_array = np.append(feat_value_array, feat_value)
    feat_num_df = pd.DataFrame(np.repeat(feat_short['NUM_FEAT'].iloc[0], feat_value_array.shape[0]),
                               columns=['NUM_FEAT'])
    feat_value_df = pd.DataFrame(feat_value_array, columns=['VALUE'])
    feat_long_df = pd.concat([feat_num_df, feat_alt_horiz_labels, feat_value_df], axis=1)
    return feat_long_df


def split(aic_long_df):
    aic_long_split_col = aic_long_df['GOF_INDEX_NUM_FEAT_SELECT'].str.split('_', 1, expand=True)
    aic_long_split_col.columns = ['GOF_INDEX', 'NUM_FEAT_SELECT']
    del aic_long_df['GOF_INDEX_NUM_FEAT_SELECT']
    aic_long_df = pd.concat([aic_long_split_col, aic_long_df], axis=1)
    return aic_long_df


def generate_long_version(feat_mic, feat_aic, feat_bic):
    mic_long_df = split(twist(feat_mic))
    aic_long_df = split(twist(feat_aic))
    bic_long_df = split(twist(feat_bic))

    mic_aic_bic_df = pd.concat([mic_long_df, aic_long_df, bic_long_df], axis=0)
    return [mic_long_df, aic_long_df, bic_long_df, mic_aic_bic_df]


def create_figures(selected_num_alt_feat, aic_long_df, bic_long_df, mic_long_df, scale):
    cut_aic_long_df = aic_long_df[aic_long_df['NUM_FEAT_ALTERED'] == selected_num_alt_feat].reset_index(drop=True)
    aic_value = cut_aic_long_df['VALUE']
    aic_num_feat_select = cut_aic_long_df['NUM_FEAT_SELECT']
    cut_bic_long_df = bic_long_df[bic_long_df['NUM_FEAT_ALTERED'] == selected_num_alt_feat].reset_index(drop=True)
    bic_value = cut_bic_long_df['VALUE']
    bic_num_feat_select = cut_bic_long_df['NUM_FEAT_SELECT']
    cut_mic_long_df = mic_long_df[mic_long_df['NUM_FEAT_ALTERED'] == selected_num_alt_feat].reset_index(drop=True)
    mic_value = cut_mic_long_df['VALUE']
    mic_num_feat_select = cut_mic_long_df['NUM_FEAT_SELECT']

    aic_bic_plt = plt.figure(figsize=(8, 6))
    plt.plot(aic_num_feat_select, aic_value, label='AIC')
    plt.plot(bic_num_feat_select, bic_value, label='BIC')
    plt.xlabel('EXTRACTED FEATURES')
    plt.ylabel('WEIGHTED VARIANCE EXPLAINED')
    plt.title('AIC/BIC FOR: ' + str(selected_num_alt_feat) + ' ALTERED FEATURES SCALED BY' + str(scale))
    plt.legend()
    plt.close()

    mic_plt = plt.figure(figsize=(8, 6))
    plt.plot(mic_num_feat_select, mic_value, label='MUTUAL INFORMATION')
    plt.xlabel('EXTRACTED FEATURES')
    plt.ylabel('TOTAL MUTUAL INFORMATION EXPLAINED')
    plt.title(
        'TOTAL MUTUAL INFORMATION EXPLAINED FOR: ' + str(selected_num_alt_feat) + ' ALTERED FEATURES SCALED BY ' + str(
            scale))
    plt.legend()
    plt.close()
    return [aic_bic_plt, mic_plt]


def find_minimum(aic_long_df, bic_long_df, selected_num_alt_feat):
    cut_aic_long_df = aic_long_df[aic_long_df['NUM_FEAT_ALTERED'] == selected_num_alt_feat].reset_index(drop=True)
    aic_min_value = np.min(cut_aic_long_df['VALUE'])
    aic_num_feat_select = int(cut_aic_long_df['NUM_FEAT_SELECT'].iloc[np.argmin(cut_aic_long_df['VALUE'])])
    aic_num_feat_select_df = pd.DataFrame([aic_num_feat_select], columns=['AIC_MIN_VALUE_NUM_FEAT_SELECT'])
    aic_min_value_df = pd.DataFrame([aic_min_value], columns=['AIC_MIN_VALUE'])

    cut_bic_long_df = bic_long_df[bic_long_df['NUM_FEAT_ALTERED'] == selected_num_alt_feat].reset_index(drop=True)
    bic_min_value = np.min(cut_bic_long_df['VALUE'])
    bic_num_feat_select = int(cut_bic_long_df['NUM_FEAT_SELECT'].iloc[np.argmin(cut_bic_long_df['VALUE'])])
    bic_num_feat_select_df = pd.DataFrame([bic_num_feat_select], columns=['BIC_MIN_VALUE_NUM_FEAT_SELECT'])
    bic_min_value_df = pd.DataFrame([bic_min_value], columns=['BIC_MIN_VALUE'])

    num_feats_alt = pd.DataFrame([selected_num_alt_feat], columns=['NUM_FEATS_ALTERED'])
    minimums_df = pd.concat(
        [num_feats_alt, aic_num_feat_select_df, aic_min_value_df, bic_num_feat_select_df, bic_min_value_df], axis=1)
    return minimums_df


def loop_figs_and_mins(feats_in_data, aic_long_df, bic_long_df, mic_long_df, scale):
    # SHOVE SCALAR IN SOMEWHERE
    full_minimums_df = pd.DataFrame([])
    full_aic_bic_plt = np.array([])
    full_mic_plt = np.array([])

    for selected_num_alt_feat in range(1, feats_in_data + 1):
        [aic_bic_plt, mic_plt] = create_figures(selected_num_alt_feat, aic_long_df, bic_long_df, mic_long_df, scale)
        full_aic_bic_plt = np.append(full_aic_bic_plt, aic_bic_plt)
        full_mic_plt = np.append(full_mic_plt, mic_plt)
        minimums_df = find_minimum(aic_long_df, bic_long_df, selected_num_alt_feat)
        full_minimums_df = pd.concat([full_minimums_df, minimums_df], axis=0)

    full_minimums_df = full_minimums_df.reset_index(drop=True)
    return [full_aic_bic_plt, full_mic_plt, full_minimums_df]


def create_output_dir(scale, selected_num_alt_feat, main_output_dir):
    if os.path.exists(main_output_dir):
        pass
    else:
        pathlib.Path(main_output_dir).mkdir(parents=True)

    os.chdir(main_output_dir)

    output_dir = main_output_dir + '/' + str(selected_num_alt_feat) + '_FEAT_ALT/' + str(scale) + '_SCALE'

    if not os.path.exists(output_dir):
        pathlib.Path(output_dir).mkdir(parents=True)
    else:
        shutil.rmtree(output_dir)
        pathlib.Path(output_dir).mkdir(parents=True)
    return output_dir


def save_figures_to_pdf(output_dir, scale, selected_num_alt_feat, full_aic_bic_plt, full_mic_plt):
    os.chdir(output_dir)

    pdf_name = 'figures_aic_bic_scale_' + str(selected_num_alt_feat) + '_' + str(scale) + '.pdf'
    pdf = matplotlib.backends.backend_pdf.PdfPages(pdf_name)
    for i in range(full_aic_bic_plt.size):
        fig = full_aic_bic_plt[i]
        pdf.savefig(fig)
    pdf.close()

    pdf_name = 'figures_mic_scale_' + str(selected_num_alt_feat) + '_' + str(scale) + '.pdf'
    pdf = matplotlib.backends.backend_pdf.PdfPages(pdf_name)
    for i in range(full_mic_plt.size):
        fig = full_mic_plt[i]
        pdf.savefig(fig)
    pdf.close()


def export_excel_workbook(output_dir, mic_long_df, aic_long_df, bic_long_df, mic_aic_bic_df, full_minimums_df, scale,
                          selected_num_alt_feat):
    os.chdir(output_dir)
    wb_name = 'mic_aic_bic_' + str(selected_num_alt_feat) + '_' + str(scale) + '.xlsx'
    wb = Workbook()
    wb.save(wb_name)
    wb = load_workbook(wb_name)
    object_array = np.array([mic_long_df, aic_long_df, bic_long_df, mic_aic_bic_df, full_minimums_df], dtype=object)
    sheet_names_array = np.array(['mic_long_df', 'aic_long_df', 'bic_long_df', 'mic_aic_bic_df', 'minimums_df'])
    for i in range(sheet_names_array.size):
        wb.create_sheet(str(sheet_names_array[i]), i)

    for i in range(sheet_names_array.size):
        active = wb[sheet_names_array[i]]

        for x in dataframe_to_rows(object_array[i], index=False):
            active.append(x)
    if 'Sheet' not in wb.sheetnames:
        pass
    else:
        ref = wb['Sheet']
        wb.remove(ref)

    wb.save(wb_name)


def export_figures_dataframes(scale, full_aic_bic_plt, full_mic_plt, mic_long_df, aic_long_df, bic_long_df,
                              mic_aic_bic_df, full_minimums_df, selected_num_alt_feat, main_output_dir):
    output_dir = create_output_dir(scale, selected_num_alt_feat, main_output_dir)
    save_figures_to_pdf(output_dir, scale, selected_num_alt_feat, full_aic_bic_plt, full_mic_plt)
    export_excel_workbook(output_dir, mic_long_df, aic_long_df, bic_long_df, mic_aic_bic_df, full_minimums_df, scale,
                          selected_num_alt_feat)


def simulate_multual_info_for_one_iteration(n, feats_in_data, scale, selected_num_alt_feat, main_output_dir):
    # RUN MAIN SCRIPT
    [_, feat_mic, feat_aic, feat_bic] = loop_mutual_info_simulation(feats_in_data, n, scale)
    # GENERATE_LONG_VERSION
    [mic_long_df, aic_long_df, bic_long_df, mic_aic_bic_df] = generate_long_version(feat_mic, feat_aic, feat_bic)
    # LOOP_FIGURE_RATION
    [full_aic_bic_plt, full_mic_plt, full_minimums_df] = loop_figs_and_mins(feats_in_data, aic_long_df, bic_long_df,
                                                                            mic_long_df, scale)
    export_figures_dataframes(scale, full_aic_bic_plt, full_mic_plt, mic_long_df, aic_long_df, bic_long_df,
                              mic_aic_bic_df, full_minimums_df, selected_num_alt_feat, main_output_dir)


def set_defaults():
    # CHANGE THESE SETTINGS TO HIGH VALUES ON A SUITABLY POWERED VM
    # [1] SET TOTAL NUMBER OF FEATURES IN DATASET
    feats_in_data = 25
    # [2] STEP VALUE
    step = 1
    # [3] SET TOTAL NUMBER OF ROWS IN DATASET
    n = 5000
    # [4] SET OUTPUT DIRECTORY
    main_output_dir = "/home/msgreen/Desktop/MI_LIKELIHOOD_SIMULATION"
    return [feats_in_data, step, n, main_output_dir]


def main():
    start_time = time.time()
    # SET SEED
    np.random.seed(123)
    [feats_in_data, step, n, main_output_dir] = set_defaults()
    for selected_num_alt_feat in range(1, feats_in_data + 1, step):
        for scale in [x / 10 for x in list(range(0, 11, 2))]:
            simulate_multual_info_for_one_iteration(n, feats_in_data, scale, selected_num_alt_feat, main_output_dir)
    sys.stdout.write(f"--- {str(time.time() - start_time)} seconds ---")


# RUN MAIN
main()
